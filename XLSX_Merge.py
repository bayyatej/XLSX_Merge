####################################################################
#    T. Bayya               XLSX_Merge.py               2/18/17    #
#------------------------------------------------------------------#
#    Copies grade data from one MSU MTH 133 Gradebook to           #
#    another. Gradebooks should be in the same directory as        #
#    the .py file.                                                 #
#    Old Gradebook name should be: 133_student_gradebook.xlsx      #
#    New gradebook name should be: 133_student_gradebook_new.xlsx  #
#------------------------------------------------------------------#
#    Link to gradebook below:                                      #
#    https://math.msu.edu/classes/mth_133/grading/Gradebook.aspx   #
####################################################################

from openpyxl import load_workbook

def get_grades(old_wb, new_wb):
    """Return 2D list of old grade cells and new grade cells as 1st and 2nd elements. Cells stored as tuples"""

    grades = [[],[]]

    old_course_grades_book = old_wb.get_sheet_names()[0]
    old_course_grades_sheet = old_wb.get_sheet_by_name(old_course_grades_book)

    old_hw_grades = old_course_grades_sheet['B2':'AI2'] #returns a 2 dimensional tuple with one element
    old_hw_grades = old_hw_grades[0]
    grades[0].append(old_hw_grades)
    old_quiz_grades = old_course_grades_sheet['B7':'M7']
    old_quiz_grades = old_quiz_grades[0]
    grades[0].append(old_quiz_grades)
    old_survey_grades = old_course_grades_sheet['R7':'S7']
    old_survey_grades = old_survey_grades[0]
    grades[0].append(old_survey_grades)
    old_exam_grades = old_course_grades_sheet['X7':'Z7']
    old_exam_grades = old_exam_grades[0]
    grades[0].append(old_exam_grades)

    new_course_grades_book = new_wb.get_sheet_names()[0]
    new_course_grades_sheet = new_wb.get_sheet_by_name(new_course_grades_book)

    new_hw_grades = new_course_grades_sheet['B2':'AI2']
    new_hw_grades = new_hw_grades[0]
    grades[1].append(new_hw_grades)
    new_quiz_grades = new_course_grades_sheet['B7':'M7']
    new_quiz_grades = new_quiz_grades[0]
    grades[1].append(new_quiz_grades)
    new_survey_grades = new_course_grades_sheet['R7':'S7']
    new_survey_grades = new_survey_grades[0]
    grades[1].append(new_survey_grades)
    new_exam_grades = new_course_grades_sheet['X7':'Z7']
    new_exam_grades = new_exam_grades[0]
    grades[1].append(new_exam_grades)

    return grades

def merge_grades(new_grades_tuple, old_grades_tuple):
    """Copies grades from old grade sheet to the new grade sheet"""
    for cell in new_grades_tuple:
        cell_index = new_grades_tuple.index(cell)
        cell.value = old_grades_tuple[cell_index].value

def main():
    old_wb = load_workbook(filename = '133_student_gradebook.xlsx')
    new_wb = load_workbook(filename = '133_student_gradebook_new.xlsx')

    grades = get_grades(old_wb, new_wb)

    new_grades_list = grades[1]
    old_grades_list = grades[0]

    for grades_tuple_index in range(len(new_grades_list)):
        merge_grades((new_grades_list[grades_tuple_index]),(old_grades_list[grades_tuple_index]))

    new_wb.save('133_student_gradebook_new.xlsx')
main()